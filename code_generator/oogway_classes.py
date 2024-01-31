"""
oogway classes
"""
import os
import re
import openpyxl
import json

class OogwayConfig:
    
    def __init__(self, oogway_templates_version, options):
        self.oogway_templates_version = oogway_templates_version
        self.options = [self.Option(**option) for option in options]

    class Option:
        def __init__(self, option, default):
            self.option = option
            self.default = default

    @classmethod
    def load(cls):
      try:
         with open("oogway_config.json", 'r') as file:
            data = json.load(file)
            return cls(**data)
      except Exception as e:
         cls.oogway_templates_version=120
         cls.options=[]
         return cls

class OogwayFun:

   @classmethod
   def welcome(cls):
    """
    Params:
      cls
    Returns: 
      nothing
   """
    # Get the directory of the script
    script_dir = os.path.dirname(os.path.realpath(__file__))

    # Construct the full path to the ASCII art file in the 'art' subfolder
    file_path = os.path.join(script_dir, 'art', 'oogway_t.ink')

    char_width=40
    lines_per_block=2
    # Open the ASCII art file and read its content with the correct encoding
    with open(file_path, 'r', encoding='utf-8') as file:
        ascii_art = file.read()

    # Split the ASCII art into lines
    lines = ascii_art.splitlines()

    # Print each block with reduced width and height
    for i in range(0, len(lines), lines_per_block):
        block = lines[i:i + lines_per_block]
        for line in block:
            print(line[:char_width])

   @classmethod
   def print_ascii_art2(cls,file_path, font_size):
      # Open the ASCII art file and read its content
      with open(file_path, 'r') as file:
         ascii_art = file.read()

      # Split the ASCII art into lines
      lines = ascii_art.splitlines()

      # Print each line with the specified font size
      for line in lines:
         print(line)
   @classmethod
   def printHelp(cls):
      script_dir = os.path.dirname(os.path.realpath(__file__))
      # Construct the full path to the ASCII art file in the 'art' subfolder
      file_path = os.path.join(script_dir, '_help', 'oogway_help.txt')
      with open(file_path, 'r') as file:
         ascii_art = file.read()

      lines = ascii_art.splitlines()

      
      for line in lines:
         print(line)

class OutputBuilder:
   sql_dialect="AWS Athena"
   orchestration="AWS Steps"
   deployment="cdk"
   code_file_extension=".json"
   platform_compute="AWS"
   platform_storage="AWS"
   platform_resource_identifier_template="arn:aws:states:[:REPLACE_AWS_REGION]:[:REPLACE_AWS_ACCOUNT]:stateMachine:[:REPLACE_RESOURCE_IDENTIFIER]"
   
   solutionName=""
   solutionLongName=""
   solutionHomePage="https://"
   auto_orchestration=1
   #manual comma separated string of database dependencies. Only used if auto orchestration = 0
   database_dependencies=""

   company_code_templates="default_company"
   company_doc_templates="default_company"

class OogwayOutputCode:
   platform_resource_identifier=""
   htnl_code_filename=""
   html_table_filename=""
   generated_code=""

class OogwayTable:
   database=""
   table_name=""
   dependencyChain=0
   #list of other oogwaytables
   dependencies=[]
   file:os.DirEntry=None
   powershell_edit=""

class OogwayCircularRefResponse:
   table_name=""
   has_circular_reference=False

class OogwayUtils:
   @classmethod
   def has_circular_reference(cls,table, visited, table_dict):
      response=OogwayCircularRefResponse()
      response.has_circular_reference=False
      
      # Mark the current table as visited
      dep_table=None
      visited.add(table.table_name)
      for dep_table_name in table.dependencies:
         for x in table_dict:
             if x==dep_table_name:
                 dep_table=table_dict[x]
                 break

         if dep_table:
            if dep_table_name in visited or cls.has_circular_reference(dep_table, visited, table_dict):
               response.has_circular_reference=True
               response.table_name=dep_table_name
               return True  # Circular reference found

      # Remove the current table from the visited set after traversal
      visited.remove(table.table_name)
      response.has_circular_reference=False
      return False  
   @classmethod
   def check_circular_references(cls,tables):
      response=OogwayCircularRefResponse()
      response.has_circular_reference=False

      # Create a dictionary to quickly look up tables by name
      table_dict = {table.table_name: table for table in tables}
      
      # Set to keep track of visited tables during traversal
      visited = set()

      # Check circular references for each table
      for table in tables:
         
         if table.table_name not in visited:
            if cls.has_circular_reference(table, visited, table_dict):
               response.has_circular_reference=True                       
               return response  # Circular reference found

      response.has_circular_reference=False
      return response

class AWSCloud:
    '''Attributes related to AWS'''
    aws_account=""
    aws_region=""
    aws_bucket=""
    aws_role_prefix=""
    aws_steps_function_role_name=""
    aws_steps_loop_name=""
    aws_db_dependency_list=""
    aws_bucket_prefix=""
    target_database=""
    raw_bucket=""
    dms_bucket=""
    staged_bucket="" 
    bucket_suffix=""
    athena_workgroup_name=""
    aws_environment=""

class IcebergOptions:
   properties=[]

class AdvOptions:
   properties=[]

class Setting:
   name=""
   value=""

class SqlHelper:
   @classmethod
   def stripCreateOrReplaceView(cls,query):
      pattern = re.compile(r"CREATE OR REPLACE VIEW .*? AS", re.IGNORECASE)
      q = re.sub(pattern, "", query)
      return q
   @classmethod
   def stripCreateOrReplaceView2(cls, query):
      """Same as stripCreateOrReplaceView but it might cover a little more aspects in the re"""
      pattern_create_view_as_with = re.compile(r"CREATE\s+OR\s+REPLACE\s+VIEW\s+\w+(\.\w+)?\s+AS", re.IGNORECASE)
      #match = re.search(pattern_create_view_as_with, sql)
      #matched_substring = match.group(1) if match else ""
      q = re.sub(pattern_create_view_as_with,"", query)
      return q
   
class ExcelHelper:
   @classmethod
   def getValueFromCell(cls, ws, cell,attrname="Cell", acceptNone=False, default=None):
        """requires an open workbook"""
        val = ws[cell].value
        if val==None:
            if acceptNone==False:
                raise Exception("Sheet:", ws.title, ", Cell:", cell, ", Message:", attrname , "Can not be Empty")
            return default
        else:
            return val
        
   @classmethod
   def getPropertyList(cls,wb, sheet, start_row, number_of_rows):
      """returns a list of Settings where name is in col A and value is in col B"""
      """params:"""
      """wb: an open workbook"""
      """sheet: name of the sheet in wb"""
      """start_row: first row in the sheet that contains the data"""
      """number_of_rows: number of rows to scan"""
      list=[]
      
      ws=wb[sheet]

      for i in range(start_row, start_row+number_of_rows):
         s=Setting()
         s.name=self.getValueFromCell(ws,'A' + str(start_row),"Name", True,"")
         s.value=self.getValueFromCell(ws,'B' + str(start_row),"Value", True,"")
         start_row=start_row+1
         list.append(s)

      return list

class DMA:
    '''Attributes related to Volvo AB DMA'''
    purpose_code_list=[]

class IcebergOptions:
    '''Attributes related to Iceberg Tables'''
    option1=""

class CRON:
    '''Attributes for CRON Schedulers'''
    year=""
    month=""
    day=""
    hour=""
    minute=""
   
class SolutionFile:
   '''
   still not in use. the idea is to use it for the main solution file. Currently it is also an instance of DvDim
   DvDim should by the way be renamed. It is basically an input table and a beeter name is requested
   '''
   fullpath:os.DirEntry
   output_builder:OutputBuilder
   output_code:OogwayOutputCode
   CRON:CRON()
   DMA:DMA()
   iceberg_options:IcebergOptions()
   
   config = {}
   
   version=0
   level=0

   aws_cloud:AWSCloud()
   #overrides (user can override the default settings)
   aws_cloud_override:AWSCloud()
   #prod values
   aws_cloud_prod:AWSCloud()
   #test values
   aws_cloud_test:AWSCloud()
   #qa values
   aws_cloud_qa:AWSCloud()
   #cdk
   aws_cloud_cdk:AWSCloud()

   def getPlatformResourceIdentifier(self,resource_name):
        '''
        Each platform has its own resource identifiers.
        This method gets the resource indetifer template from the solution.xlsx file, replaces some tags (if any)
        and finaly returns the platform resource identifier.
        '''
        rid= self.output_builder.platform_resource_identifier_template
        
        
        rid=rid.replace("[:REPLACE_AWS_ACCOUNT_NUMBER]",self.aws_cloud.aws_account)
        rid=rid.replace("[:REPLACE_AWS_REGION]",self.aws_cloud.aws_region)
        rid=rid.replace("[:REPLACE_RESOURCE_IDENTIFIER]",resource_name)

        return rid
   
   def checkVersion(self):
        is_valid=True
        wb = openpyxl.load_workbook(self.fullpath, data_only = True, read_only=True)
        try:
            ws = wb['Version']
            self.version=ws['B1'].value
            if self.version==None or self.version<=103:
                  wb.close()
                  is_valid=False
                  Util.print(WARNING,"Unsuported version for ", self.fullpath.path, ". Version is: ", str(self.version))
               
        except Exception as e:
           raise e
           
        finally:
            wb.close()

        return is_valid 
        
   def  __init__(self,fullpath: os.DirEntry):
        #Util.print(INFO, fullpath.path)
        self.fullpath=fullpath
        self.output_code=OogwayOutputCode()
        self.level=0
        self.version=0
        self.checkVersion()