import json
from math import fabs
from operator import is_
from pathlib import Path
from posixpath import split
import re
from sre_constants import SUCCESS
import string
from sys import prefix
from time import process_time_ns
from tokenize import group
import openpyxl
import glob
import os
import json
from datetime import datetime
from oogway_classes import IcebergOptions, DMA,OogwayOutputCode, OogwayTable, OogwayUtils, AWSCloud, AdvOptions,IcebergOptions, Setting, SqlHelper, CRON
import sqlparse
#https://pypi.org/project/sql_metadata/
from sql_metadata import Parser

#from emworks_AWSStepsBuilder import AWSStepsBuilder
#from emworks_AWSStepsBuilder import SQL

from emworks_templates import *
import shutil
from Util import Util, INFO,DEBUG1,DEBUG2,DEBUG3,WARNING,ERROR,SOLUTIONERROR, HINT
import io
import subprocess
import uuid
import time
from oogway_classes import OutputBuilder

global g_DROP_EXISTING_TABLE
g_DROP_EXISTING_TABLE=True

"""SolutionConfig"""
class SolutionConfig(object):
    '''
    parses an excel file and saves the cell values for the defined name in excel.
    Only single value cells are allowed, no cell ranges.
    '''
    substitution_tags = {}
    files_with_path = []
    def __init__(
            self,
            excel_input_folder:str,
            tag_prefix:str = 'replace_'
            #path_to_file='/home/a236868/repo/odl/code_generation_framework/type1_raw_to_staged/excel_templates',
            #file_name='solution.xlsx'
        ) -> dict:
        tags = self.parse_excel_folder(excel_input_folder, tag_prefix)

        self.substitution_tags = {**tags, **self.substitution_tags}

    def parse_excel_file(
            self,
            path_to_file: str,
            file_name: str,
            tag_prefix: str) -> dict:
        '''
        Read an excel file and convert its defined name values to simple key value paisrs in a dictionary.
        Returns this dictionary.
        only load keys (tags) that start with the string defined in tag_prefix
        '''
        config = {}
        file_with_path = os.path.join(path_to_file, file_name)
        self.files_with_path.append(file_with_path)
        work_book = openpyxl.load_workbook(file_with_path, data_only=True)
        for sheet_name in work_book.sheetnames:
            if " " in sheet_name:
                Util.print(ERROR,f"Workbook may not have sheet names with spaces. {work_book.sheetnames}")
                raise Exception("Error: Sheet names may not contain space")

        defined_names_dict = work_book.defined_names
        for key, value in defined_names_dict.items():
            if tag_prefix in key:
                #print(f'key, Cell position, value: {key}, {value.attr_text}, {value.value}')
                sheet_name = value.attr_text.split('!')[0]
                excel_sheet = work_book[sheet_name]
                cell_defined_name = value.attr_text.split('!')[1]
                #print(f"cell_defined_name: {cell_defined_name}")
                try:
                    cell_value = excel_sheet[cell_defined_name].value
                except Exception as e:
                    print(f"Error for {key}, {cell_defined_name}")
                    raise e
                #print(f'excel defined name, excel defined value: {key}, {cell_value}')
                config[key] = cell_value
                if cell_value is None:
                    print(f"Warning, empty key value for {key}")
        if "solution" not in file_name.lower():
            root_name = config["REPLACE_TARGET_TABLE"]
        else:
            root_name = "GLOBAL"
        return {root_name:config}
    def parse_excel_folder(self, excel_input_folder: str, tag_prefix: str):
        '''
        Add key value paris to the config dictionary for this code generation "solution".
        '''

        tags = {}
        for excel_file in os.scandir(excel_input_folder):
            if "~$" != excel_file.name[0:2] and ".xlsx" == excel_file.name[-5:]:
                print(f"Reading code generator input file: {excel_file.name}")
                tags = {
                    **tags,
                    **self.parse_excel_file(
                        path_to_file=excel_input_folder,
                        file_name=excel_file.name,
                        tag_prefix=tag_prefix)
                }
            else:
                print(f"skipping file:{excel_file.name}")
        return tags

class configFile:
    '''COnfig file for oogway. Still not in use'''
    solutionName=""
    tables=[]

class SolutionPaths:
    '''Paths and files used by oogway'''
    company_code_folder="default_company"
    company_html_folder="default_company"

    root_folder=""
    main_path=Path.cwd()
    latest_oogway_version=122

    documentation_templates_path=""

    #this will be the full path to the solution.xlsx file during build...
    solutionfile_build_fullpath=""

    mod_path=Path(__file__).parent
    oogway_home_rel_path=""
    oogway_home_powershell_scripts_rel_path=""
    oogway_code_templates_folder="code_templates"
    solution_file_full_path=""
    oogway_config_filename="solutionconfig.json"

    oogway_template_rel_path=""

    #oogway directory structure
    oogway_root_path=""
    oogway_build_path=""
    oogway_folder="oogway"
    oogway_build_folder="build"
    oogway_temp_folder="tmp"
    oogway_upgrade_folder="upgrade"
    oogway_upgrade_path=""

    oogway_table_template="oogway_table_template"
    oogway_sql="oogway_sql"

    oogway_solution_template_file_name=""
    oogway_solution_file_name="solution.xlsx"

    #path to the solution file in the solution folder
    oogway_solution_file_path=""

    oogway_home_solution_file_rel_path=""
    oogway_home_documentation_path=""
    oogway_home_bash_helper_scripts_path=""
    oogway_backup_folder="backup"
    oogway_backup_path=""

    #this is the folder where the signed update_excel_file.ps1 script should be located
    oogway_powershell_folder="powershell"
    oogway_powershell_path=""

    latest_oogway_sql_full_path=""
    powershell_helper_scripts_path=""
    bash_helper_scripts_path=""
    oogway_documentation_path=""
    latest_oogway_template=""
    template_cdk_path=""
    template_cdk_fullpath = ""
    doctemplate_rel_path= ""
    code_template_rel_path = ""
    old_files_folder=""
    code_template_path = ""
    oogway_template_path= ""
    oogway_sql_full_path = ""

    excel_input_folder=""
    code_output_root_path:string = ""
    doc_output_path = ""
    code_output_path=""
    aws_lambda_output_path = ""
    aws_cloud_bridge_output_path = ""
    aws_roles_output_path = ""
    aws_cdk_output_path = ""

    oogway_excel_templates_folder="excel_templates"
    oogway_powershell_helper_scripts_folder="powershell_helper_scripts"
    oogway_bash_helper_scripts_folder="bash_helper_scripts"
    oogway_documentation_templates_folder = 'documentation_templates'

    def openConfigFile():
        pass

    def load_json_file(file_path):
        try:
            with open(file_path, 'r') as json_file:
                data = json.load(json_file)
                return data
        except FileNotFoundError:
            print(f"File not found: {file_path}")
            return None
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
            return None

    def saveConfigFile():
        pass

    def updateExcelFileWithPowershell(self, target_fullpath: string):

        folder = self.oogway_powershell_path
        file = os.path.join(folder,"update_excel_file.ps1")
        p=subprocess.Popen(["powershell.exe",file,"-fullpath " + target_fullpath ],stdout=subprocess.PIPE)
        p.communicate()

    def set_root_folder(self,root_folder:string):
        self.root_folder=root_folder

        self.oogway_solution_template_file_name="solution_v" +str(self.latest_oogway_version) + ".xlsx"
        self.oogway_root_path=os.path.join(self.root_folder, self.oogway_folder)
        self.oogway_build_path=os.path.join(self.oogway_root_path,self.oogway_build_folder)
        self.oogway_temp_path=os.path.join(self.oogway_root_path,self.oogway_temp_folder)
        self.oogway_upgrade_path=os.path.join(self.oogway_root_path,self.oogway_upgrade_folder)

        self.oogway_home_rel_path=".."
        #where to find excel templates
        self.oogway_template_rel_path = os.path.join(self.oogway_home_rel_path ,"templates", self.oogway_excel_templates_folder)

        #where to find the powershell helper scripts...
        
        self.oogway_home_powershell_scripts_rel_path=os.path.join(".." ,"shell_scripts", "powershell_helper_scripts")
        self.oogway_home_powershell_signed_scripts_rel_path = os.path.join(".." ,"shell_scripts", "powershell_signed_scripts")
        self.oogway_home_bash_helper_scripts_path= os.path.join(".." ,"shell_scripts", "bash_helper_scripts")

        #solution powershell helper scripts
        
        self.powershell_helper_scripts_path = os.path.join(self.root_folder,self.oogway_powershell_helper_scripts_folder)
        self.oogway_powershell_path=os.path.join(self.oogway_root_path, self.oogway_powershell_folder)
        self.bash_helper_scripts_path=os.path.join(self.root_folder,self.oogway_bash_helper_scripts_folder)

        self.oogway_home_solution_file_rel_path=os.path.join(self.oogway_template_rel_path, self.oogway_solution_template_file_name)

        #sql engine
        self.latest_oogway_sql= self.oogway_sql + "_v" + str(self.latest_oogway_version) + ".xlsx"
        self.latest_oogway_sql_full_path=os.path.join(self.oogway_template_rel_path,self.latest_oogway_sql)

        #table template file
        self.latest_oogway_table_template= self.oogway_table_template + "_v" + str(self.latest_oogway_version) + ".xlsx"
        self.oogway_table_template_full_path=os.path.join(self.oogway_template_rel_path,self.latest_oogway_table_template)

        self.oogway_home_templates_path=os.path.join(self.oogway_home_rel_path,"templates", "excel_templates")
        self.oogway_home_documentation_path=os.path.join(self.oogway_home_rel_path,"documentation","html")
        self.oogway_documentation_path=os.path.join(self.oogway_root_path,"documentation")

        self.oogway_backup_path=os.path.join(self.oogway_root_path,self.oogway_backup_folder)

        self.excel_input_folder=os.path.join(root_folder,'database')
        self.solution_file_full_path=os.path.join(self.excel_input_folder,"solution.xlsx")

        self.code_output_root_path = os.path.join(root_folder , "deploy")
        self.doc_output_path = os.path.join(root_folder,"doc","html")

        self.aws_lambda_output_path = self.code_output_root_path + r'\lambda'
        self.aws_cloud_bridge_output_path = self.code_output_root_path + r'\event_bridge'
        self.aws_roles_output_path = self.code_output_root_path + r'\roles'
        self.aws_cdk_output_path = self.code_output_root_path + r'\cdk'
        self.aws_roles_output_path = self.code_output_root_path + r'\roles'

        self.oogway_solution_file_path=os.path.join(self.excel_input_folder,self.oogway_solution_file_name)
        self.oogway_sql_full_path = os.path.join(self.oogway_template_rel_path, self.latest_oogway_sql)

    def add_table(self,table_to_add:string):
        src = self.oogway_table_template_full_path

        trg_filename = table_to_add + "_v" + str(self.latest_oogway_version) + ".xlsx"
        trg = os.path.join(self.oogway_build_path,trg_filename)
        final_trg = os.path.join(self.excel_input_folder ,trg_filename)

        if(os.path.isfile(final_trg)):
            Util.print (WARNING,"The table " + table_to_add + " already exists in the solution.")
        else:
            Util.print(INFO,"Adding file ", trg, " to solution...")
           
            Util.copy(src,trg)
            src_wb=openpyxl.load_workbook(trg,data_only=False, read_only=False)
            ws=src_wb['Options']
            ws['B21']=table_to_add
            src_wb.save(trg)
            self.updateExcelFileWithPowershell(trg)
            
            Util.copy(trg,final_trg)

            Util.print(INFO,"Successfully added the table ", table_to_add, " in file ", trg, " to the solution")

    def useupgrade(self):
        self.backup()
        count_upgrade = self.countfiles(self.oogway_upgrade_path)
        Util.print(INFO,"Number of files in upgrade folder:", str(count_upgrade))
        count_current = self.countfiles(self.excel_input_folder)
        Util.print(INFO,"Number of files in input folder:", str(count_current))
        if count_current==count_upgrade:
            Util.print(INFO, "Clear input path", self.excel_input_folder)
            Util.clearDir(self.excel_input_folder)

            #print("Copy files from ", self.oogway_upgrade_path, " to ", self.excel_input_folder)
            shutil.copytree(self.oogway_upgrade_path,self.excel_input_folder, dirs_exist_ok=True)
            #print("Clear upgrade path: ", self.oogway_upgrade_path)
            Util.clearDir(self.oogway_upgrade_path)
            Util.print(INFO,"Files successfully moved from upgrade to: " , self.excel_input_folder)

    def countfiles(self,dir_path:string):
        count=0
        Util.print(INFO,"Counting files in path:", dir_path)
        for path in os.listdir(dir_path):
            # check if current path is a file
            if os.path.isfile(os.path.join(dir_path, path)):
                count += 1
        return count

    def backup(self):

        radom_folder = str(uuid.uuid4())
        backup_path = os.path.join(self.oogway_backup_path,radom_folder)

        Util.print(INFO,"Backup files to:", backup_path)

        Util.mkdir(backup_path)
        for f in os.scandir(self.excel_input_folder):
            #f is a DirEntry
            if "~$" != f.name[0:2] and ".xlsx" == f.name[-5:]:
                target_file = os.path.join(backup_path, f.name)
                Util.copy(f.path, target_file)

    def init_solution(self, solution_name:string, refresh_only:bool):
        
        #powershell_helper_scripts
        src = self.oogway_home_powershell_scripts_rel_path
        trg = self.powershell_helper_scripts_path
        Util.copyTree(src,trg)

        #powershell signed scrips
        src=self.oogway_home_powershell_signed_scripts_rel_path
        trg=self.oogway_powershell_path
        Util.copyTree(src,trg)

        #bash helper scripts
        src=self.oogway_home_bash_helper_scripts_path
        trg=self.bash_helper_scripts_path
        Util.copyTree(src,trg)

        #oogway documentation
        src=self.oogway_home_documentation_path
        trg=self.oogway_documentation_path
        Util.copyTree(src,trg)


        if refresh_only==False:
            if(os.path.exists(self.solution_file_full_path)):
                Util.print(INFO,"Solution.xlsx is already in place. Nothing is done with it...")
            else:

                    src = self.oogway_home_solution_file_rel_path

                    trg1 = os.path.join(self.oogway_build_path,"solution.xlsx")
                    trg2 = self.oogway_solution_file_path

                    Util.print(DEBUG2, "Copying ", src , "to ", trg1)
                    shutil.copy(src,trg1)
                    if(solution_name!=""):
                        Util.print(DEBUG2,"Update ", trg1, " Set solution name to ", solution_name)
                        Util.print(DEBUG2, "Loading excel file ", trg1 )
                        src_wb=openpyxl.load_workbook(trg1,data_only=False, read_only=False)
                        ws=src_wb['Options']
                        ws['B2']=solution_name
                        Util.print(DEBUG2, "Saving excel file ", trg1 )
                        src_wb.save(trg1)
                        src_wb.close()
                        self.updateExcelFileWithPowershell(trg1)
                        Util.print("Copying ", trg1 , " to ", trg2)
                        shutil.copy(trg1,trg2)

    def debugprint(self,logger):

        Util.print(INFO,"Working directory: " + self.excel_input_folder )

    def createfolders(self, clear_folders:bool=True):
        if os.path.exists(self.root_folder)==False:
            Util.mkdir(self.root_folder)
        if os.path.exists(self.excel_input_folder)==False:
            Util.mkdir(self.excel_input_folder)

        Util.mkdir(self.oogway_root_path) #\oogway
        Util.mkdir(self.oogway_build_path) #\oogway\build
        Util.mkdir(self.oogway_temp_path) #oogway\tmp
        Util.mkdir(self.oogway_upgrade_path) #\oogway\upgrade

        Util.mkdir(self.oogway_powershell_path) #\oogway\powershell
        #powershell scripts
        src = self.oogway_home_powershell_scripts_rel_path
        trg = self.oogway_powershell_path

        if os.path.exists(trg)==False:
            shutil.copytree(src,trg, dirs_exist_ok=True)
        else:

            Util.clearDir(trg)
            shutil.copytree(src,trg,dirs_exist_ok= True)



        #Util.mkdir(self.oogway_old_files_path) #\oogway\old
        #Util.mkdir(self.powershell_helper_scripts_path)

        Util.mkdir(self.bash_helper_scripts_path)

        Util.mkdir(self.doc_output_path) #\doc\html
        images = os.path.join(self.doc_output_path,"images")
        Util.mkdir(images)
        Util.mkdir(self.code_output_root_path)
        Util.mkdir(self.oogway_backup_path)
        Util.clearDir(self.oogway_temp_path)

        Util.clearDir(self.oogway_build_path)
        if clear_folders:
            Util.clearDir(self.oogway_upgrade_path)
            Util.clearDir(self.doc_output_path)
            Util.clearDir(self.oogway_build_path)
            #Util.clearDir(self.oogway_old_files_path)

    def copyImages(self):

        src_logo = os.path.join( self.documentation_templates_path, "images", "logo.png")
        trg_logo = os.path.join(self.doc_output_path,"images")
        Util.print(DEBUG1,"Copy images from ", src_logo, " to ", trg_logo)
        #Util.mkdir(trg_logo)
        trg_logo=os.path.join(trg_logo,"logo.png")

        Util.copy(src_logo,trg_logo)

    def setCode(self, output_builder: OutputBuilder):

        code = output_builder.orchestration
        self.company_code_folder= output_builder.company_code_templates
        self.company_html_folder= output_builder.company_doc_templates



        Util.print(DEBUG1,"Set Code:", code)
        #code: i.e "steps","dlt"
        #where to find the code templates
        self.code_template_rel_path = os.path.join('..'
                                                   ,"templates"
                                                   ,self.oogway_code_templates_folder
                                                   ,self.company_code_folder
                                                   , code)

        self.code_template_path = os.path.join(self.main_path,self.code_template_rel_path)

        #where to write the output code
        self.code_output_path=os.path.join(self.code_output_root_path,code)
        Util.clearDir(self.code_output_path)

        self.template_cdk_path=os.path.join(".."
                                            ,"templates"
                                            ,self.oogway_code_templates_folder
                                            ,self.company_code_folder
                                            , "cdk")

        self.template_cdk_fullpath = str((self.main_path/self.template_cdk_path).resolve() )

        self.doctemplate_rel_path= os.path.join('..'
                                                ,"templates"
                                                ,self.oogway_documentation_templates_folder
                                                ,self.company_html_folder
                                                , code
                                                , 'html')
        self.documentation_templates_path=str((self.main_path/self.doctemplate_rel_path).resolve() )

        Util.mkdir(self.code_output_path)
        #util.copy()
        self.copyImages()

class CDKHelper:
    '''CDKHelper copies CDK related files to the solution repo'''
    version="1"
    comment=""
    template_cdk_fullpath=""
    cdk_output_root_path=""


    solutionName=""
    deploy_to=""
    bucket=""
    prefix=""
    fullpath=""
    aws_athena_workgroup=""
    aws_region=""
    databasename=""

    cdk_stackname_with_underscore=""
    cdk_stack_classname=""
    cdk_stack_classname_with_dash=""
    cdk_infra_stackname_with_dash=""

    comma_separated_stacknames_with_dash=""
    comma_separated_stacknames_with_dash_prod=""

    step_functions_role_name=""

    aws_role_prefix=""

    aws_region_qa_start_with_dash=""
    aws_region_prod_start_with_dash=""

    cron=CRON()

    #path "cdk_custom_stacks"


    #root
    deploy_files=[

            {"src":"app.py", "dst":"app.py", "folder":""},
            {"src":"cdk.json","dst":"cdk.json","folder":""},
            #{"src":"configuration-prod.json","dst":"configuration-prod.json","folder":""},
            #{"src":"configuration-qa.json","dst":"configuration-qa.json","folder":""},
            {"src":".gitignore","dst":".gitignore","folder":""},

            {"src":"projectspec.yaml","dst":"projectspec.yaml","folder":""},
            {"src":"requirements.txt","dst":"requirements.txt","folder":""},
            {"src":"requirements-dev.txt","dst":"requirements-dev.txt","folder":""},

            {"src":"__init__.py", "dst":"__init__.py", "folder":"cdk_custom_stacks"},
            {"src":"env_variables.py", "dst":"env_variables.py", "folder":"cdk_custom_stacks"},
            {"src":"infra_stack.py", "dst":"infra_stack.py", "folder":"cdk_custom_stacks"},
            {"src":"steps_stack.py", "dst":"steps_stack.py", "folder":"cdk_custom_stacks"},
            {"src":"generic_stack.py","dst":"generic_stack.py","folder":"cdk_custom_stacks"}
    ]

    def __init__(self,cdk_output_root_path, template_root_path, main_solution_file):

        self.cdk_output_root_path=cdk_output_root_path
        self.template_cdk_fullpath=template_root_path
        #self.fullpath=fullpath

        self.comment=""


        self.solutionName=main_solution_file.output_builder.solutionName
        self.cdk_stackname_with_underscore=self.solutionName + "_stack"
        self.cdk_stack_classname=self.solutionName + "Stack"
        self.cdk_stack_classname_with_dash=self.solutionName + "-stack"
        self.cdk_infra_stackname_with_dash="infra-stack"

        #prod, qa or cdk
        self.deploy_to = main_solution_file.deploy_to

        self.aws_role_prefix=main_solution_file.aws_cloud.aws_role_prefix
        self.step_functions_role_name=main_solution_file.aws_cloud.aws_steps_function_role_name
        self.bucket=main_solution_file.aws_cloud.aws_bucket
        self.prefix=main_solution_file.aws_cloud.aws_bucket_prefix
        self.fullpath="s3://" + self.bucket + "/" + self.prefix

        Util.print(INFO,"Target Database:", main_solution_file.target_database)
        self.databasename=main_solution_file.target_database

        self.aws_athena_workgroup=main_solution_file.aws_cloud.athena_workgroup_name

        self.environment=main_solution_file.aws_cloud.aws_environment
        self.aws_account=main_solution_file.aws_cloud.aws_account
        self.aws_region= main_solution_file.aws_cloud.aws_region

        self.fullpath="s3://" + self.bucket + "/" + self.prefix

        self.comment=main_solution_file.comment
        if(self.comment==None):
            self.comment=""



        role_db_dependencies = main_solution_file.aws_cloud.aws_db_dependency_list
        Util.print(DEBUG1,"Role DB Dependencies:", role_db_dependencies)
        self.step_role_db_dependencies=role_db_dependencies

        self.cron=main_solution_file.cron



        #derived attributes
        if self.deploy_to=="qa":
            self.aws_environment_qa="qa"
            self.aws_environment_prod=""
            self.aws_region_qa_start_with_dash="-" + self.aws_region
            self.aws_region_prod_start_with_dash=""
        elif self.deploy_to=="prod":
            self.aws_environment_qa=""
            self.aws_environment_prod="prod"
            self.aws_region_prod_start_with_dash="-" + self.aws_region
            self.aws_region_qa_start_with_dash=""
        elif self.deploy_to=="cdk":
            self.aws_environment_qa="qa"
            self.aws_environment_prod="prod"
            self.aws_region_prod_start_with_dash= "- eu-west-1"


        #output paths for CDK
        cdk_output_common_code=os.path.join(self.cdk_output_root_path,"cdk_custom_stacks")

        Util.mkdir(cdk_output_common_code)
        Util.clearDir(cdk_output_common_code)

    def upgrade(self, to_version, oogway_template_file_full_path):

        if (self.version==to_version):
            Util.print(INFO,"\tAlready latest version")
        else:
             dst = str((self.fullpath / "name.xls" ).resolve() )

             #logger.debug("Copy to " + self.fullpath.name + "_2")
             shutil.copy(oogway_template_file_full_path,dst)

    def copyToDeploy(self):

        current_date = Util.getCurrentDate()

        for s in self.deploy_files:

            dst_path=os.path.join(self.cdk_output_root_path,s["folder"])
            src_path=os.path.join(self.template_cdk_fullpath,s["folder"])

            src=os.path.join(src_path ,  s["src"])

            s["dst"]=s["dst"].replace("[:REPLACE_SOLUTION_NAME]", self.solutionName)

            dst=os.path.join(dst_path, s["dst"])


            shutil.copy(src,dst)

            with open(src,'r') as src_file:
                data = src_file.read()
                src_file.close()
                with open(dst,'w+') as file:
                    data=data.replace("[:REPLACE_COMMENT]",self.comment)
                    data=data.replace("[:REPLACE_SOLUTION_NAME]",self.solutionName)
                    data=data.replace("[:REPLACE_CURRENT_DATE]",current_date)
                    data=data.replace("[:REPLACE_CDK_STACKNAME_CLASSNAME]",self.cdk_stack_classname)
                    data=data.replace("[:REPLACE_TARGET_BUCKET]",self.bucket)
                    data=data.replace("[:REPLACE_AWS_ACCOUNT_NUMBER]",self.aws_account)

                    data=data.replace("[:REPLACE_AWS_REGION]",self.aws_region)

                    data=data.replace("[:REPLACE_CDK_STACKNAME_WITH_DASH]",self.cdk_stack_classname_with_dash)
                    data=data.replace("[:REPLACE_CDK_STACKNAME_WITH_UNDERSCORE]",self.cdk_stackname_with_underscore)
                    data=data.replace("[:REPLACE_CDK_INFRA_STACKNAME_WITH_DASH]",self.cdk_infra_stackname_with_dash)

                    data=data.replace("[:REPLACE_DATABASE_NAME]",self.databasename)
                    data=data.replace("[:REPLACE_ATHENA_WORKGROUP]",self.aws_athena_workgroup)



                    data=data.replace("[:REPLACE_COMMA_SEPARATED_STACKNAMES_WITH_DASH_PROD]",self.comma_separated_stacknames_with_dash_prod)
                    data=data.replace("[:REPLACE_COMMA_SEPARATED_STACKNAMES_WITH_DASH]",self.comma_separated_stacknames_with_dash)


                    data=data.replace("[:REPLACE_STEP_FUNCTIONS_ROLE_NAME]",self.step_functions_role_name)

                    data=data.replace("[:REPLACE_ODL_ENVIRONMENT_QA_LOWER_CASE]",self.aws_environment_qa)
                    data=data.replace("[:REPLACE_ODL_ENVIRONMENT_PROD_LOWER_CASE]",self.aws_environment_prod)

                    data=data.replace("[:REPLACE_AWS_REGION_QA_STARTS_WITH_DASH]", self.aws_region_qa_start_with_dash)
                    data=data.replace("[:REPLACE_AWS_REGION_PROD_STARTS_WITH_DASH]", self.aws_region_prod_start_with_dash)

                    data=data.replace("[:REPLACE_ROLE_PREFIX]", self.aws_role_prefix)


                    data=data.replace("[:REPLACE_ODL_ENVIRONMENT_PROD_LOWE_CASE]","prod:")
                    data=data.replace("[:REPLACE_ODL_ENVIRONMENT_QA_LOWE_CASE]","qa:")

                    data=data.replace("[:REPLACE_CDK_INFRA_STACK_NAME_WITH_DASH]",self.cdk_infra_stackname_with_dash)
                    data=data.replace("[:REPLACE_ROLE_DB_DEPENDENCIES]",self.step_role_db_dependencies)

                    file.write(data)

class DvDim():
    '''DvDim is actually a table. TODO: Rename properly'''
    version=0
    DMA:DMA()
    cron:CRON #move to main solution
    #effective values
    aws_cloud:AWSCloud()
    #overrides (user can override the default settings)
    aws_cloud_override:AWSCloud()
    #prod values
    aws_cloud_prod:AWSCloud()
    #test values
    aws_cloud_test:AWSCloud()
    #qa values
    aws_cloud_qa:AWSCloud()
    #cdk
    aws_cloud_cdk:AWSCloud()

    #iceberg options
    iceberg_options:IcebergOptions()

    #adv options
    adv_options:AdvOptions()

    output_builder: OutputBuilder
    output_code:OogwayOutputCode
    config = {}
    table_name=""
    table_name_mini_dimension=""
    target_database=""

    #main solution files (TODO: Move)
    deploy_to=""
    deploy_to_prev=""

    resource_name=""
    resource_name_mini_dimension=""
    output_code=""

    doc_hmtl_filename=""
    #unload_table_steps_name=""
    dependencyChain=[]

    level=0

    dep_tables=[]
    dep_tables_source_view=[]

    #outgoing views (end user views)
    list_of_view_names=[]

    #incoming views (used for loading the tables)
    list_of_incoming_view_names=[]

    dep_tables_other=[]

    #The source tables for any custom views created for the end users...
    dep_tables_views=[]

    depchain=[]

    #list of tables in the same use case this table has dependencies to
    target_table_dependencies=[]

    method=1
    load_method=1
    is_valid=True
    fullpath=""

    #tmp
    list_of_incoming_views=[]
    list_of_views=[]
    list_of_views_as_string=""



    list_of_incoming_views_as_string=""

    original_list_of_views_as_string=""
    original_list_of_incoming_views_as_string=""

    def  __init__(self,fullpath: os.DirEntry):
        #Util.print(INFO, fullpath.path)
        self.fullpath=fullpath
        self.output_code=OogwayOutputCode()
        self.level=0
        self.version=0
        self.checkVersion()

    def replaceTags(self, val):
        for key in self.config:
            replace_tag="[:" + key + "]"

            val=val.replace(replace_tag, str(self.config[key]))
        return val

    def checkVersion(self):
        wb = openpyxl.load_workbook(self.fullpath, data_only = True, read_only=True)
        ws = wb['Version']
        self.version=ws['B1'].value
        if self.version==None or self.version<=103:
            wb.close()
            self.is_valid=False
            Util.print(WARNING,"Unsuported version for ", self.fullpath.path, ". Version is: ", str(self.version))

        else:
            ws = wb['Options']
            self.table_name=ws['B22'].value

        wb.close()

        #if self.table_name==None:
            #Util.print(ERROR,"Table name is missing in file:", self.fullpath.path)
            #sys.exit(1)

        return self.table_name

    def formatSql2(self, sql, parseTables:bool, html_table_of_tables:string):
        html_table=""
        html_rows=""

        list=[]
        if(sql==""):
            return []

        if(parseTables==True):
            #print(sql)
            html_table="<table><th>Table Name</th>"
            list = Parser(sql).tables

            try:
                filtered_list = [item for item in list if "." in item]
                for t in filtered_list:
                    html_rows=html_rows + "<tr><td>" + t +'</td></tr>'
            except Exception as e:
                print(e)

            html_table=html_table+ html_rows + "</table>"

        sql=sqlparse.format(sql, reindent=True,keyword_case='upper')
        return list, sql, html_table

    def formatSql(self, configTag:string, parseTables:bool, html_table_of_tables:string):
        has_warning=False
        html_table=""
        html_rows=""
        sql_to_parse=""
        sql_formatted=""



        configTagFormatted=configTag+"_FORMATTED"
        self.config[configTagFormatted]=""

        list=[]

        sql = self.getConfigValue(configTag,'')
        Util.print(DEBUG3,sql)
        sql = self.replaceTags(sql)

        self.config[configTag]=sql
        sql_to_parse=sql
        #pattern_cross_join = re.compile(r'CROSS JOIN .*?(\s|$)', re.IGNORECASE)
        sql_to_parse = SqlHelper.stripCreateOrReplaceView2(sql)

        Util.print(DEBUG3,"Trying to sql format configTag:", configTag, "parseTables:", str(parseTables))

        if  "cross join unnest" in sql.lower():
            Util.print(WARNING, "The select statement uses a cross join unnest.. ")
            has_warning=True
        elif "unnest(" in sql.lower():
            Util.print(WARNING, "The select statement uses a join unnest.. ")
            has_warning=True

        if(sql==""):
            return []

        list,sql_formatted,html_table = self.formatSql2(sql_to_parse,parseTables,html_table_of_tables)
        self.config[html_table_of_tables]=html_table
        self.config[configTagFormatted]=sql_formatted

        filtered_list=list
        if has_warning:
            filtered_list = [item for item in list if "." in item]
            Util.print(DEBUG1,str(filtered_list))


        return filtered_list

    def getValueFromCell(self, ws, cell,attrname="Cell", acceptNone=False, default=None):
        val = ws[cell].value
        if val==None:
            if acceptNone==False:
                raise Exception("Sheet:", ws.title, ", Cell:", cell, ", Message:", attrname , "Can not be Empty")
            else:
                Util.print(DEBUG3,"Sheet", ws.title, ", Cell:", cell,"Attribute:", attrname, " is empty..." )
            return default
        else:
            return val

    def reset_deploy_to(self):
        if self.reset_deploy_to!="":
            wb = openpyxl.load_workbook(self.fullpath, data_only = False, read_only=False)
            try:
                Util.print(INFO,"Resetting deploy to: ", self.deploy_to_prev)
                if self.version>120:
                    sheet="Options"
                    cell="B3"
                else:
                    sheet="AWS"
                    cell="G3"

                ws = wb[sheet]
                ws[cell].value=self.deploy_to_prev


                wb.save(self.fullpath)
            except Exception as e:
                raise e
            finally:
                wb.close()

    def set_deploy_to(self,solution_paths, val:string, action):

        solution_paths.solutionfile_build_fullpath = os.path.join(solution_paths.oogway_build_path,self.fullpath.name)

        if val!="" and action=="build":
            wb = openpyxl.load_workbook(self.fullpath, data_only = False, read_only=False)
            try:
                Util.print(INFO,"Updating deploy to: ", val)
                if self.version<120:
                    sheet="Options"
                    cell="B3"

                elif self.version>=120:
                    sheet="AWS"
                    cell="G3"

                ws = wb[sheet]
                deploy_to_prev=self.getValueFromCell(ws,cell, "deploy to")
                ws[cell].value=val

                wb.save(solution_paths.solutionfile_build_fullpath)

                solution_paths.updateExcelFileWithPowershell(solution_paths.solutionfile_build_fullpath)

            except PermissionError:
                Util.print(ERROR, "Permission error on file ", self.fullpath, " . Make sure you do not have it opened in Excel!")
                sys.exit(1)
            except Exception as e:
                raise e
            finally:
                wb.close()
        else:
            #not a build action, then we only need to copy the file... (faster)
            Util.copy(self.fullpath, solution_paths.solutionfile_build_fullpath)

    def _getPropertyList(self,wb, sheet, start_row, number_of_rows):
        list=[]

        ws=self.getSheet(wb,sheet)
        for i in range(start_row, start_row+number_of_rows):
            s=Setting()
            s.name=self.getValueFromCell(ws,'A' + str(start_row),"Name", True,"")
            s.value=self.getValueFromCell(ws,'B' + str(start_row),"Value", True,"")
            start_row=start_row+1
            list.append(s)

        return list

    def _readSheetOptions(self,wb, version):
        #OPTIONS
        Util.print(DEBUG1,"Reading sheet options")
        ws = self.getSheet(wb,"Options")
        if version<120:

            self.output_builder.solutionName=self.getValueFromCell(ws,"B2", "Solution Name")
            self.output_builder.solutionLongName=self.getValueFromCell(ws,"B8", "Solution Long Name")
            self.comment=self.getValueFromCell(ws,"B26", "Comment",True,"")
            self.output_builder.solutionHomePage=self.getValueFromCell(ws,"B9", "Solution Home Page")
            self.output_builder.database_dependencies=self.getValueFromCell(ws,"B31", "Database Dependencies List")

        elif version>=120:

            self.output_builder.solutionName=self.getValueFromCell(ws,'B2',"Solution Name")
            self.output_builder.solutionLongName=self.getValueFromCell( ws, 'B3','Solution Long Name')
            self.output_builder.solutionHomePage=self.getValueFromCell( ws, 'B4','Home Page')
            self.comment=self.getValueFromCell(ws,"B6", "Comment", True)
            self.output_builder.database_dependencies=self.getValueFromCell(ws,"B7","Database Dependencies")

            if version>120:
                self.output_builder.company_code_templates =self.getValueFromCell(ws,"B9", "Code Template Company", True,"default_company")
                self.output_builder.company_doc_templates=self.getValueFromCell(ws,"B10", "Doc Template Company", True,"default_company")

    def _readSheetAWS(self,wb,version):
        Util.print(DEBUG1,"Reading sheet AWS")
        self.aws_cloud=AWSCloud()
        self.aws_cloud_override=AWSCloud()
        #AWS CLOUD ONLY EFFECTIVE VALUES IS READ (Column F)
        if version<120:
            ws=self.getSheet(wb,"Options")
            self.deploy_to=self.getValueFromCell(ws,"B3", "Deploy To")
            self.aws_cloud.aws_bucket=self.getValueFromCell(ws,"B4", "AWS Bucket Name")
            self.aws_cloud.aws_bucket_prefix=self.getValueFromCell(ws,"B5", "AWS Bucket Prefix")
            self.aws_cloud.athena_workgroup_name=self.getValueFromCell(ws,"B7", "AWS Athena Workgroup Name")
            self.aws_cloud.aws_environment=self.getValueFromCell(ws,"E12", "AWS Environment")
            self.aws_cloud.aws_account=self.getValueFromCell(ws,"E13", "AWS Account")
            self.aws_cloud.aws_region=self.getValueFromCell(ws,"E14", "AWS Region")
            self.aws_cloud.raw_bucket=self.getValueFromCell(ws,"E15", "AWS RAW Ingestion Bucket")
            self.aws_cloud.dms_bucket=self.getValueFromCell(ws,"E16", "AWS DMS Ingestion Bucket")
            self.aws_cloud.staged_bucket=self.getValueFromCell(ws,"E17", "AWS Staged Bucket")
            self.aws_cloud.bucket_suffix=self.getValueFromCell(ws,"E22", "AWS Bucket Name Suffix")
            self.aws_cloud.aws_role_prefix=self.getValueFromCell(ws,"B37","AWS Role Prefix")
            self.aws_cloud.aws_steps_function_role_name=self.getValueFromCell(ws,"B28","AWS Steps Function Role Name")

            self.target_database=self.getValueFromCell(ws,"B6", "AWS Athena Database Name")

        elif version>=120:
            ws = self.getSheet(wb,"AWS")
            self.aws_cloud.aws_account=self.getValueFromCell(ws,"F4", "AWS Account")
            self.aws_cloud.aws_region=self.getValueFromCell(ws,"F5", "AWS Region")
            self.aws_cloud.raw_bucket=self.getValueFromCell(ws,"F6", "RAW Bucket")
            self.aws_cloud.dms_bucket=self.getValueFromCell(ws,"F7", "DMS Bucket")
            self.aws_cloud.dms_bucket=self.getValueFromCell(ws,"F8", "Staged Bucket")
            self.aws_cloud.bucket_suffix=self.getValueFromCell(ws,"F9", "Bucket Suffix")
            self.aws_cloud.aws_role_prefix=self.getValueFromCell(ws,"F10", "Role  Prefix")
            self.aws_cloud.aws_steps_function_role_name=self.getValueFromCell(ws,"F11", "Step Function Role Name")
            self.aws_cloud.aws_bucket=self.getValueFromCell(ws,"F12", "Bucket Name")
            self.aws_cloud.aws_bucket_prefix=self.getValueFromCell(ws,"F13", "Prefix Name")

            self.target_database= self.getValueFromCell(ws,'F14', "AWS Target Database Name")



            self.aws_cloud.athena_workgroup_name = self.getValueFromCell(ws,"F15", "AWS Athena Workgroup Name")

            self.deploy_to = self.getValueFromCell(ws,'G3', "Deploy To")
            self.aws_cloud.aws_environment=self.getValueFromCell(ws,'G3', "AWS Environment")

    def _readSheetOutput(self,wb,version):
        Util.print(DEBUG1,"Reading sheet OUTPUT")

        #OUTPUT --- WARNING ----
        ws=wb["OUTPUT"] #TODO: Never read directly from OUTPUT
        self.list_of_views_as_string =self.getValueFromCell( ws,'B29',True,"[]")
        self.list_of_incoming_views_as_string=self.getValueFromCell(ws,'B35',True,"[]")

        #remove new lines
        self.list_of_views_as_string=self.list_of_views_as_string.replace("\n", " ").replace("\r", " ")# .replace('"', '')
        self.list_of_incoming_views_as_string = self.list_of_incoming_views_as_string.replace("\n", " ").replace("\r", " ") #.replace('"', '')

        #those view will later be create view....
        self.list_of_incoming_views = json.loads(self.list_of_incoming_views_as_string)
        self.list_of_views = json.loads(self.list_of_views_as_string)

        #original
        self.original_list_of_incoming_views_as_string=self.list_of_incoming_views_as_string
        self.original_list_of_views_as_string=self.list_of_views_as_string

    def _readSheetCRON(self,wb, version):
        sheet="CRON"
        Util.print(DEBUG1,"Reading sheet ", sheet)

        ws=wb[sheet]
        self.cron.year=self.getValueFromCell( ws,'A4',"year")
        self.cron.month=self.getValueFromCell( ws,'B4',"month")
        self.cron.day=self.getValueFromCell( ws,'C4',"day")
        self.cron.hour=self.getValueFromCell( ws,'D4',"hour")
        self.cron.minute=self.getValueFromCell( ws,'E4',"minute")

    def _readSheetDMA(self,wb, version):
        sheet="DMA"
        Util.print(DEBUG1,"Reading sheet ", sheet)
        ws=self.getSheet(wb,sheet)
        self.DMA.purpose_code_list.append(self.getValueFromCell(ws, 'A2',"Purpose Code 1",True,""))
        self.DMA.purpose_code_list.append(self.getValueFromCell(ws, 'A3',"Purpose Code 2",True,""))
        self.DMA.purpose_code_list.append(self.getValueFromCell(ws, 'A4',"Purpose Code 3",True,""))
        self.DMA.purpose_code_list.append(self.getValueFromCell(ws, 'A5',"Purpose Code 4",True,""))
        self.DMA.purpose_code_list.append(self.getValueFromCell(ws, 'A6',"Purpose Code 5",True,""))

        self.DMA.purpose_code_list = [item for item in self.DMA.purpose_code_list if item != ""]

        Util.print(DEBUG1,"DMA Purpose Code List:", str(self.DMA.purpose_code_list))

    def _readSheetIcebergOptions(self,wb, version):
        self.iceberg_options=IcebergOptions()
        #Iceberg Options
        self.iceberg_options.properties = self._getPropertyList(wb,"IcebergOptions",1,10)

    def _readSheetAdvOptions(self,wb,version):
         #AdvOptions
        self.adv_options=AdvOptions()
        self.adv_options.properties = self._getPropertyList(wb,"AdvOptions",2,6)

    def _readSheetCode(self,wb,version):
        #CODE
        if version<120:
            sheet="Options"
            ws=wb[sheet]
            self.output_code.sql_dialect = self.getValueFromCell(ws,"B35", "Sql Dialect")
            Util.print(DEBUG1,"Sql Dialect:", self.output_builder.sql_dialect)

            self.output_builder.orchestration=self.getValueFromCell(ws,"B36", "Orchestration")

            Util.print(DEBUG1,"Orchestration:", self.output_builder.orchestration)

            self.output_builder.deployment=self.getValueFromCell(ws,"B37", "Sql Dialect")
            self.output_builder.code_file_extension=self.getValueFromCell(ws,"B38", "Code File Extension")
            self.output_builder.platform_compute=self.getValueFromCell(ws,"B39", "Platform Compute")
            self.output_builder.platform_storage=self.getValueFromCell(ws,"B40", "Platform Storage")
            self.output_builder.auto_orchestration=self.getValueFromCell(ws,"B41", "Auto Orchestration")


        if version>=120:
            ws = wb["Code"]
            self.output_builder.sql_dialect=self.getValueFromCell(ws,'B2', "SQL Dialect")
            self.output_builder.orchestration=self.getValueFromCell( ws,'B3',True,"aws_steps")
            self.output_builder.deployment=self.getValueFromCell( ws,'B4',"Deployment")
            self.output_builder.code_file_extension=self.getValueFromCell( ws,'B5', "Code File Extension")
            self.output_builder.platform_compute=self.getValueFromCell( ws,'B6', "Compute Platform")
            self.output_builder.platform_storage=self.getValueFromCell( ws,'B7',"Storage Platform")
            self.output_builder.platform_resource_identifier_template=self.getValueFromCell(ws, 'B8', "platform resource identifier template")
            self.output_builder.auto_orchestration=self.getValueFromCell(ws,"B9", "Auto Orchestration")

    def _writeSheetOptions(self,new_wb):
        '''Upgrading main_soltion_file'''

        ws=self.getSheet(new_wb,"Options")
        ws["B2"].value= self.output_builder.solutionName
        ws["B3"].value=self.output_builder.solutionLongName
        ws["B4"].value=self.output_builder.solutionHomePage

        ws["B6"].value = self.comment
        ws["B7"].value=self.output_builder.database_dependencies

        ws["B9"].value=self.output_builder.company_code_templates
        ws["B10"].value=self.output_builder.company_doc_templates

    def _writeSheetCode(self, new_wb):
        #Code generation settings
        ws=self.getSheet(new_wb,"Code")
        self.setCellValue(ws,"B2", self.output_builder.sql_dialect,"Sql Dialect")
        self.setCellValue(ws,"B3", self.output_builder.orchestration,"Orchestration")
        self.setCellValue(ws,"B4", self.output_builder.deployment,"Deployment")
        self.setCellValue(ws,"B5", self.output_builder.code_file_extension,"Code File Extension")
        self.setCellValue(ws,"B6", self.output_builder.platform_compute,"Platform Compute")
        self.setCellValue(ws,"B7", self.output_builder.platform_storage,"Platform Storage")
        self.setCellValue(ws,"B8", self.output_builder.platform_resource_identifier_template,"Resource Indentifier Template")
        self.setCellValue(ws,"B9", self.output_builder.auto_orchestration,"Auto Orchestration")

    def readMainSolution(self, solution_paths, oogway_py_version, action):
        self.output_builder=OutputBuilder()
        self.cron=CRON()
        self.aws_cloud=AWSCloud()
        self.DMA=DMA()
        self.resource_name="main"

        Util.print(DEBUG1,"Open workbook:", self.fullpath)
        wb = openpyxl.load_workbook(self.fullpath, data_only = True, read_only=True)
        try:

            #VERSION
            Util.print(DEBUG1,"Reading Version")
            ws=self.getSheet(wb,"Version")
            self.version=self.getValueFromCell(ws,"B1","Version")
            version=self.version
            Util.print(DEBUG1,"\tSoulution.xlsx has version:", str(self.version))

            if(action!="upgrade"):
                if(self.version!=oogway_py_version):
                    Util.print(WARNING, "\tConsider upgrading your Excel files to the latest version ")
                    Util.print(HINT,"\t\tUse the action -a upgrade to do this!")

            self._readSheetOptions(wb, version)
            self._readSheetAWS(wb,version)
            self._readSheetCode(wb,version)
            self._readSheetOutput(wb,version)
            self._readSheetCRON(wb,version)
            self._readSheetDMA(wb,version)
            self._readSheetIcebergOptions(wb,version)
            self._readSheetAdvOptions(wb,version)


        except Exception as e:
            Util.print(SOLUTIONERROR, e)
        finally:
            wb.close()


        self.dep_tables_views=[]
        self.list_of_view_names=[]

        #otgoing view analysis:
        self.scanViews(self.list_of_views, "REPLACE_CUSTOM_VIEWS_HTML_TABLE", "outgoing views")
        #incoiming views
        self.scanViews(self.list_of_incoming_views, "REPLACE_INCOMING_VIEWS_HTML_TABLE", "incoming views")



        Util.print (INFO,"\tBuilding: ", self.output_builder.orchestration)
        Util.print (INFO,"\tSql Dialect: ", self.output_builder.sql_dialect)
        Util.print (INFO,"\tDeployment: ", self.output_builder.deployment)


    def check_table_format(self,array):
        for item in array:
            if '.' not in item:  # Check if there's at least one dot
                return False
            
            parts = item.split('.')
            if any(part.strip() == '' for part in parts):  # Check if there's text before and after each dot
                return False

        return True

    def scanViews(self,list_of_views,tag,name):
        """TODO: Are there any tables that are not fully quailified? If yes then model error"""
        view_html_row=""
        _view_list=[]

        number_of_views=len(list_of_views)
        if number_of_views>0:

            Util.print(INFO,"\tThere are ", str(number_of_views), " ", name ,"  view(s) in the solution.")

            for v in list_of_views:
                _view = OogwayTable()
                q_value=SqlHelper.stripCreateOrReplaceView(v.get('q', ''))

                _view.table_name = self.target_database + "." + v.get('name','')
                a=[]
                a, sql, html_list_of_tables =self.formatSql2(q_value,True,"[:UNUSED1]")
                _view.dependencies=a

                if self.check_table_format(a)==False:
                    Util.print(HINT,"The view with name ", v.get('name',''), " references an unqualified table. Please correct this")
                    Util.print(HINT, "\tSource Tables:", str(a)) 
                    Util.print(HINT, "All views needs to use fully qualified source tables. If you are referencing a table in the same database then you could use [:REPLACE_TARGET_DATABASE_NAME] as database name.")   
                    Util.print(HINT, "All views are defined in solution.xlsx")
                    Util.print(SOLUTIONERROR, "Sorry about this. Please fix and retry!")
                result_string = ", ".join(a)

                

                view_html_row = view_html_row + "<tr><td>" + _view.table_name + "</td><td>" + sql + "</td><td>" + result_string + "</td></tr>"

                self.dep_tables_views=list(set(self.dep_tables_views+a))



                self.list_of_view_names.append(_view.table_name)

                _view_list.append(_view)

            view_html_row="<table><th>Name</th><th>Select Statement</th><th>Source Tables Used</th>" + view_html_row + "</table>"

            self.config[tag]=view_html_row

            Util.print(DEBUG1,"Source tables used by the views:", str(self.dep_tables_views))


            response = OogwayUtils.check_circular_references(_view_list)
            if response.has_circular_reference:
                Util.print(SOLUTIONERROR,"The ", name, " has at least one circular dependency . Please check " , response.table_name ,"  in the solution.xlsx file.")

    def openForBuild(self,main_solution_file, fullpath):

        html_variables="<table>"
        html_variable=""

        #READS ALL VARIABLES FROM THE EXCEL SHEET
        self.fullpath=fullpath
        Util.print(DEBUG1, "Open build file ", self.fullpath, " to get the [:REPLACE_....] variables...")
        wb = openpyxl.load_workbook(self.fullpath, data_only = True, read_only=True)
        defined_names_dict = wb.defined_names
        for key, value in defined_names_dict.items():

            if(key[0:8])=="REPLACE_":

                sheet_name = value.attr_text.split('!')[0]
                excel_sheet = wb[sheet_name]
                cell_defined_name = value.attr_text.split('!')[1]
                try:

                    computed_value =  excel_sheet[cell_defined_name].value
                    cell_value = excel_sheet[cell_defined_name].value
                    if(cell_value==None):
                        cell_value="<Not Set>"

                    html_variable="<tr><td>" + key + "</td><td>" + str(cell_value) +"</td></tr>"

                except Exception as e:
                    Util.print(ERROR,f"Error for {key}, {cell_defined_name}")
                    raise e

                self.config[key] = cell_value

        wb.close()
        #FIX SOME AND ADD SOME EXTRA VARIABLES
        try:
            self.load_method=self.config["REPLACE_TEMPLATE_NUMBER"]
            self.target_database=main_solution_file.target_database

            #format sql
            #new variables will be created with the same name but _FORMATTED added in the end.
            #i.e for the variable REPLACE_SOURCE_VIEW the new REPLACE_SOURCE_VIEW_FORMATTED will be added.
            #use the _FORMATTED in html and the original in your code templates. The original sql is always an
            #unformatted one line sql string.
            self.formatSql('REPLACE_SOURCE_VIEW', False,'')
            self.formatSql('REPLACE_CREATE_TARGET_TABLE', False,'')

            l1=[]

            #if self.load_method==3 or self.load_method==2:
            dimensions = self.getConfigValue("REPLACE_LIST_OF_QUALIFIED_DIMENSIONS","")
            l1 = dimensions.split(", ")



            

            #self.formatSql('REPLACE_SQL_CUSTOM_SOURCE_VIEW',False)
            Util.print(DEBUG1,"REPLACE_SQL_CUSTOM_SOURCE_VIEW")
            l2=self.formatSql('REPLACE_SQL_CUSTOM_SOURCE_VIEW',True,'REPLACE_CUSTOM_SOURCE_VIEW_TABLES')

            self.dep_tables=list(set(l1+l2))
            #print(str(self.dep_tables))



            Util.print(DEBUG1,"REPLACE_STAGE_VIEW")
            self.formatSql('REPLACE_STAGE_VIEW',False,'')


            Util.print(DEBUG3,str(self.dep_tables))

            self.formatSql('REPLACE_TEST_SOURCE_BK',False,'')
            self.config['REPLACE_SOLUTION_LONG_NAME'] = main_solution_file.output_builder.solutionLongName
            self.config['REPLACE_CODE_GENERATOR_ORCHESTRATION']=main_solution_file.output_builder.orchestration
            use_scd4 = self.getConfigValue("REPLACE_USE_MINI_DIMENSION",0)
            self.getConfigValue("REPLACE_SHOW_HTML_MINI_DIMENSION","display: none")

            if  str.lower(self.getConfigValue("REPLACE_SOURCE_TABLE_VISABILITY","hidden"))=='hidden':
                self.config['REPLACE_SOURCE_TABLE_VISABILITY']='display: none'
                self.config['REPLACE_SOURCE_TABLE_DISPLAY']='display: none'
                self.config['REPLACE_SOURCE_VIEW_DISPLAY']='display: block'

            else:
                self.config['REPLACE_SOURCE_TABLE_VISABILITY']='display: block'
                self.config['REPLACE_SOURCE_VIEW_DISPLAY']='display: none'
                self.config['REPLACE_SOURCE_TABLE_DISPLAY']='display: block'

            #dependency chain is if the table has any dependencies to other tables loaded by this job.
            #a 2 means that the table will be loaded in group #2 after dim and facts
            #self.dependencyChain=self.getConfigValue("REPLACE_DEPENDENCY_CHAIN",1)

        except Exception as e:
            Util.print(ERROR,e)


        self.version=self.config["REPLACE_CODEGEN_VERSION"]
        self.aws_account=self.getConfigValue("REPLACE_AWS_ACCOUNT_NUMBER","unknown")
        self.aws_region=self.getConfigValue("REPLACE_AWS_REGION","unknown")

        self.resource_name=self.getConfigValue("REPLACE_RESOURCE_IDENTIFIER","main")

        self.resource_name_mini_dimension=self.getConfigValue("REPLACE_MINI_DIMENSION_RESOURCE_NAME","")
        self.aws_steps_loop_name=self.getConfigValue("REPLACE_AWS_STEPS_LOOP_NAME","[:REPLACE_AWS_STEPS_LOOP_NAME] did not exist!")
        self.doc_hmtl_filename=self.config["REPLACE_HTML_DOC_FILENAME"]
        self.table_name=self.getConfigValue("REPLACE_TARGET_TABLE_NAME","")
        self.table_name_mini_dimension=self.getConfigValue("REPLACE_TABLE_NAME_MINI_DIMENSION","")

        if(self.table_name==None):
            print("ERROR: No table name!")
            self.table_name=""

        current_date = Util.getCurrentDate()
        self.config["REPLACE_CURRENT_DATE"]=current_date

    def getConfigValue(self,tag: string, defaultValue:string=None):
        try:
            val = self.config[tag]
            if(val==None or val=='<Not Set>'):
                  val=defaultValue
                  self.config[tag]=defaultValue
            return val
        except :
            if(defaultValue!=None):
                self.config[tag]=defaultValue
                return defaultValue
            else:
                e= "Tag:" + tag + ", could not be found and no default value was provided!"
                raise Exception(e)

    def updateFromSolution(self, solution_paths:SolutionPaths, target_fullpath:string, main_solution_file):
        '''copies data from the solution.xlsx file to the build file'''

        output_builder = OutputBuilder()
        Util.print(DEBUG1,"Merging solution.xlsx and meta data")

        #open target file
        Util.print(DEBUG1,"\tOpen ", target_fullpath, " for read")
        trg_wb = openpyxl.load_workbook(target_fullpath, read_only=False)


        trg_ws = trg_wb['Options']
        trg_ws['D2'] = main_solution_file.output_builder.solutionName
        trg_ws['D4'] = main_solution_file.target_database
        trg_ws['D23'] = main_solution_file.target_database
        trg_ws['D28'] = main_solution_file.aws_cloud.aws_bucket
        trg_ws['D29'] = main_solution_file.aws_cloud.aws_bucket_prefix
        trg_ws['D67'] = main_solution_file.aws_cloud.aws_account
        trg_ws['D68'] = main_solution_file.aws_cloud.aws_region
        trg_ws['D69'] = main_solution_file.aws_cloud.athena_workgroup_name
        trg_ws['D70'] = main_solution_file.target_database


        #output builder
        output_builder=main_solution_file.output_builder


        #DMA
        start_row=36
        for index, p in enumerate(main_solution_file.DMA.purpose_code_list):
            cell='D' + str(start_row)
            Util.print(DEBUG2,"\tPurpose Code:", p, " is written to cell ", cell)
            trg_ws[cell] = p
            start_row=start_row+1



        #AdvOptions
        start_row=43
        for index, s in enumerate(main_solution_file.adv_options.properties):
            cell='D' + str(start_row)
            Util.print(DEBUG2,"\Advance Option:", s.value, " is written to cell ", cell)
            trg_ws[cell] = s.value
            start_row=start_row+1


        #Iceberg Options
        start_row=52
        end_row=62
        for index, s in enumerate(main_solution_file.iceberg_options.properties):
            trg_ws['D' + str(start_row)] = s.value
            start_row=start_row+1


        trg_wb.save(target_fullpath)
        trg_wb.close()

        return output_builder

    def prepareBuild(self, solution_paths:SolutionPaths, source_filename: string, main_solution_file):
        source_filename = source_filename
        target_fullpath=os.path.join(solution_paths.oogway_build_path,source_filename)

        Util.copy(solution_paths.latest_oogway_sql_full_path,target_fullpath)
        self.updateFromSolution(solution_paths,target_fullpath, main_solution_file)

        #input data
        src = os.path.join(solution_paths.excel_input_folder,self.fullpath.name)
        Util.print (INFO,"\tOpen input file :",src, "for read" )
        src_wb = openpyxl.load_workbook(src, data_only = True, read_only=True)

        Util.print(INFO, "\tOpen build file : [:yellow]",target_fullpath, "[:default] for write." )
        trg_wb = openpyxl.load_workbook(target_fullpath, data_only=False, read_only=False)



        #logger.debug("\tCopy data from input file to build file...")
        #here is the copying work performed
        #sheet: options (done)
        #sheet: doc  (done)
        #sheet: custom source_view (done)
        #sheet: columns
        #sheet: dimensions

        src_ws = src_wb['Options']
        trg_ws = trg_wb['Options']
        #template number
        trg_ws['B7'] = src_ws['B4'].value

        #source database
        trg_ws['B10'] = src_ws['B7'].value

        #source table
        trg_ws['B11'] = src_ws['B8'].value

        #load date col name
        trg_ws['B13'] = src_ws['B10'].value

        #last modified
        trg_ws['B14'] = src_ws['B11'].value

        #last modified data type
        trg_ws['B15'] = src_ws['B12'].value

        #distinct values
        trg_ws['B17'] = src_ws['B14'].value

        #cusom source view
        trg_ws['B19'] = src_ws['B16'].value

        #disable table prefix
        trg_ws['B21'] = src_ws['B18'].value

        #logical table name
        trg_ws['B24'] = src_ws['B21'].value

        #use scd-2
        trg_ws['B32'] = src_ws['B29'].value

        #use dma
        trg_ws['B34'] = src_ws['B31'].value
        #vehiclprofileid column name
        trg_ws['B35'] = src_ws['B32'].value


        ws="Custom Source View"
        src_ws = src_wb[ws]
        trg_ws = trg_wb[ws]
        trg_ws['A2'] = src_ws['A2'].value

        ws="Doc"
        try:
            src_ws = src_wb[ws]
        except:
            ws="confluence"
            src_ws = src_wb[ws]

        trg_ws = trg_wb[ws]
        trg_ws['B2'] = src_ws['B2'].value
        trg_ws['B6'] = src_ws['B6'].value

        ws="COLUMNS"
        src_ws = src_wb[ws]
        trg_ws = trg_wb[ws]
        for r in range(2, 150):
            for c in range(1,15):
                #note the common o for both, could also be o+1 for one if there is an offset
                trg_ws.cell(r, c).value = src_ws.cell(r, c).value

        ws="DIMENSIONS"
        src_ws = src_wb[ws]
        trg_ws = trg_wb[ws]
        for r in range(4, 16):
            for c in range(1,10):
                #note the common o for both, could also be o+1 for one if there is an offset
                trg_ws.cell(r, c).value = src_ws.cell(r, c).value





        trg_ws=trg_wb["OUTPUT"]


        defined_names_dict = trg_wb.defined_names

        for key, value in defined_names_dict.items():

            if(key[0:8])=="REPLACE_":

                sheet_name = value.attr_text.split('!')[0]
                excel_sheet = trg_wb[sheet_name]
                cell_defined_name = value.attr_text.split('!')[1]
                try:

                    computed_value =  excel_sheet[cell_defined_name].value
                    cell_value = excel_sheet[cell_defined_name].value

                    #print( key, " = ", cell_value)
                except Exception as e:
                    print(f"Error for {key}, {cell_defined_name}")
                    raise e

                self.config[key] = cell_value

                #print("key",key,"=",cell_value )



        trg_wb.save(target_fullpath)
        trg_wb.close()
        src_wb.close()

        return target_fullpath
    #builds the ouput files
    def build(self,solution_paths,source_file_name:string, main_solution_file):

        target_fullpath = self.prepareBuild(solution_paths,source_file_name, main_solution_file)

        #openpyxl can not recalculte the excel formulas
        #it only uses a cached version that was calculated last time excel opened the file.
        #please check kola2 python lib that seems to handle this issue
        #now we need to open the files in excel and save them
        Util.print (INFO,"\tRefreshing ", target_fullpath, " using powershell and Excel. " )

        solution_paths.updateExcelFileWithPowershell(target_fullpath)


        self.openForBuild(main_solution_file, target_fullpath)

    def setCellValue(self,ws,cell, value, comment):
        try:
            ws[cell].value=value
        except Exception as e:
            Util.print(ERROR, e)

    def getSheet(self,wb,sheet_name):
        try:
            ws=wb[sheet_name]
        except Exception as e:
            Util.print(ERROR, e)
        return ws

    def _setAWSCloud(self,wb, group_name, letter, aws: AWSCloud):
        #AWS Overides (Overrides default excel formulas)
        ws=wb["AWS"]
        self.setCellValue(ws, letter + "4", aws.aws_account,"AWS Account " + group_name)
        self.setCellValue(ws,letter + "5", aws.aws_region,"AWS Region " + group_name)
        self.setCellValue(ws,letter+ "6", aws.raw_bucket,"AWS Raw Bucket " + group_name)
        self.setCellValue(ws,letter+"7", aws.dms_bucket,"AWS DMS Bucket " + group_name)
        self.setCellValue(ws,letter+"8", aws.staged_bucket,"AWS Staged Bucket " + group_name)
        self.setCellValue(ws,letter+"9", aws.bucket_suffix,"AWS Bucket Suffix " + group_name)

        self.setCellValue(ws,letter+"10",aws.aws_role_prefix,"AWS Role Prefix " + group_name)
        self.setCellValue(ws,letter+"11",aws.aws_steps_function_role_name,"AWS Step Function Role Name " + group_name)
        self.setCellValue(ws,letter+"12", aws.aws_bucket, "AWS Bucket " + group_name)
        self.setCellValue(ws,letter+"13", aws.aws_bucket_prefix,"AWS Bucket Prefix " + group_name)
        self.setCellValue(ws,letter+"14",aws.target_database, "Target Database Name " + group_name)
        self.setCellValue(ws,letter+"15", aws.athena_workgroup_name, "Athena Workgroup Name " + group_name)

    def _copyNewSolutionFile(self,solution_paths):
        original_src= self.fullpath
        new_template =  solution_paths.oogway_solution_template_file_name

        src = os.path.join( solution_paths.excel_input_folder, "solution.xlsx")
        trg = os.path.join(solution_paths.oogway_build_path,"solution.xlsx")
        Util.print(DEBUG1, "\tCopy ", src, " to ", trg)
        try:
            shutil.copy(src,trg)

        except Exception as e:
            Util.print(ERROR, e)

        return trg

    def upgradeMainSolutionFile(self, solution_paths):

        trg = self._copyNewSolutionFile(solution_paths)

        new_template_soure = os.path.join(solution_paths.oogway_home_templates_path,solution_paths.oogway_solution_template_file_name)
        new_template_trg = os.path.join(solution_paths.oogway_build_path,solution_paths.oogway_solution_template_file_name)

        Util.print(DEBUG1,"\tCopy ", new_template_soure, " to ", new_template_trg)
        try:
            shutil.copy(new_template_soure,new_template_trg)
        except PermissionError as pe:
            Util.print(ERROR,"Could not copy to ", new_template_trg, ". If you have this file opened in Excel then please close it and try again." )
            sys.exit(1)

        target_fullpath=os.path.join(solution_paths.oogway_build_path,"solution.xlsx")
        Util.print(DEBUG1,"\tOpen new solution file ", new_template_trg, " for write!")
        new_wb = openpyxl.load_workbook(new_template_trg, data_only = False, read_only=False)
        try:

            #Options
            self._writeSheetOptions(new_wb)
            #Code
            self._writeSheetCode(new_wb)



            #AWS Settings Supports Prod, Test, QA and CDK (all have default values in solution.xlsx, override is if user wants other values)
            self._setAWSCloud(new_wb,"Override", "G", self.aws_cloud_override)
            #self._setAWSCloud(new_wb,"Prod","B", self.aws_cloud_prod)
            #self._setAWSCloud(new_wb,"Test","C", self.aws_cloud_test)
            #self._setAWSCloud(new_wb,"QA", "D", self.aws_cloud_qa)
            #self._setAWSCloud(new_wb,"CDK", "E", self.aws_cloud_cdk)



            #DMA
            ws = new_wb["DMA"]
            purpose_codes = self.DMA.purpose_code_list
            for index, p in enumerate(purpose_codes):
                row=index+2
                cell="A" + str(row)
                self.setCellValue(ws,cell, p,"Purpose Code " + str(index+1))

            #CRON
            ws = new_wb["CRON"]
            self.setCellValue(ws,"A4", self.cron.year,"CRON Year")
            self.setCellValue(ws,"B4", self.cron.month,"CRON Month")
            self.setCellValue(ws,"C4", self.cron.day,"CRON Day")
            self.setCellValue(ws,"D4", self.cron.hour,"CRON Hour")
            self.setCellValue(ws,"E4", self.cron.minute,"CRON Minute")



            #outgoing views: TODO: Remove the create or replace view
            ws = new_wb["Views"]
            for index, v in enumerate(self.list_of_views):
                i=index+2
                q_value = SqlHelper.stripCreateOrReplaceView(v.get('q', ''))
                self.setCellValue(ws,"B" + str(i), v["name"], "Outgoing View Name " + str(index+1))
                self.setCellValue(ws,"C" + str(i), q_value, "Outgoing View Select " + str(index+1))



            #incoming views: TODO Remove the create or replace views
            ws = new_wb["InView"]
            for index, v in enumerate(self.list_of_incoming_views):
                i=index+2
                q_value = SqlHelper.stripCreateOrReplaceView(v.get('q', ''))
                self.setCellValue(ws,"B" + str(i), v["name"], "Incoming View Name " + str(index+1))
                self.setCellValue(ws,"C" + str(i), q_value, "Incoming View Select " + str(index+1))




            Util.print(DEBUG1,"Saving ", new_template_trg)
            new_wb.save(new_template_trg)
        except Exception as e:
            Util.print(ERROR, e)

        finally:
            new_wb.close()

        solution_paths.updateExcelFileWithPowershell(new_template_trg)
        dest = os.path.join( solution_paths.oogway_upgrade_path,"solution.xlsx")
        Util.copy(new_template_trg,dest)





    #upgrade
    def upgrade(self, solution_paths:SolutionPaths, name: string):
        upgrade_needed=False

        #the user file (from database)
        original_src = os.path.join(solution_paths.excel_input_folder,self.fullpath.name)
        #this is the new template copied from oogway template folder
        original_trg = os.path.join(solution_paths.oogway_build_path, self.fullpath.name)

        Util.print(DEBUG1,"\tCopy existing file :",original_src, " to ", original_trg )
        shutil.copy(original_src,original_trg)
        Util.print (DEBUG1, "\tRefresh original file ", original_trg)
        solution_paths.updateExcelFileWithPowershell(original_trg)

        Util.print(DEBUG1, "\tOpen ", original_trg, " for read")
        original_src_wb = openpyxl.load_workbook(original_trg, data_only = False, read_only=True)
        ws=original_src_wb['Version']
        self.version=ws['B1'].value
        Util.print(DEBUG1,"\tVersion:", self.version)

        if(self.version==solution_paths.latest_oogway_version):
            target_fullpath=os.path.join(solution_paths.oogway_build_path,self.fullpath.name)
            shutil.copy(original_src,target_fullpath)

            Util.print(DEBUG1, "\tNo upgrade needed. Already latest version.")
            return
        elif(self.version<=104):
            upgrade_needed=True
            ws=original_src_wb['Options']
            self.table_name=ws['B24'].value

        elif(self.version>=105 <120):
            ws=original_src_wb['Options']
            self.target_database=""
            self.table_name=ws['B21'].value
        elif (self.version>120):
            ws=original_src_wb['Options']
            self.target_database=""
            self.table_name=ws['B22'].value

        Util.print(DEBUG1, "\tTable Name:", self.table_name)


        #update to latest table template

        #this is where the upgraded file will be found
        new_target_file = self.table_name + "_v" + str(solution_paths.latest_oogway_version) +  ".xlsx"
        final_dest = os.path.join(solution_paths.oogway_upgrade_path, new_target_file)
        target_fullpath = os.path.join(solution_paths.oogway_build_path,new_target_file)

        new_target_fullpath = os.path.join(solution_paths.excel_input_folder,new_target_file)


        oogway_tample_template_file = solution_paths.oogway_table_template_full_path

        Util.print(DEBUG1,"\tCopying template ", oogway_tample_template_file, " to ", target_fullpath)
        Util.copy(oogway_tample_template_file,target_fullpath)
        Util.print(DEBUG1, "\tOpen file :",target_fullpath, "for write" )
        trg_wb = openpyxl.load_workbook(target_fullpath)

        #print("\tCopy data from old file to new file...")
        #here is the copying work performed
        #sheet: options (done)
        #sheet: doc  (done)
        #sheet: custom source_view (done)
        #sheet: columns
        #sheet: dimensions

        src_ws = original_src_wb['Options']
        trg_ws = trg_wb['Options']

        if(self.version<=104):
            #template number
            trg_ws['B4'] = src_ws['B7'].value
            #source database
            trg_ws['B7'] = src_ws['B10'].value
            #source table/view
            trg_ws['B8'] = src_ws['B11'].value
            #load date column name
            trg_ws['B10'] = src_ws['B13'].value
            #last_modified_column name
            trg_ws['B11'] = src_ws['B14'].value
             #last_modified_column data type
            trg_ws['B12'] = src_ws['B15'].value
            #distinct values only
            trg_ws['B14'] = src_ws['B17'].value
            #custom source view
            trg_ws['B16'] = src_ws['B19'].value
            #disable table prefix
            trg_ws['B18'] = src_ws['B21'].value
            #logical table name
            trg_ws['B21'] = src_ws['B24'].value
            #is dimension table
            trg_ws['B28'] = src_ws['B31'].value
            #use scd-2
            trg_ws['B29'] = src_ws['B32'].value
            #use dma
            trg_ws['B31'] = src_ws['B33'].value
            #vehicleplatformid column name
            trg_ws['B32'] = src_ws['B34'].value
        else:
             #template number
            trg_ws['B4'] = src_ws['B4'].value
            #source database
            trg_ws['B7'] = src_ws['B7'].value
            #source table/view
            trg_ws['B8'] = src_ws['B8'].value
            #load date column name
            trg_ws['B10'] = src_ws['B10'].value
            #last_modified_column name
            trg_ws['B11'] = src_ws['B11'].value
             #last_modified_column data type
            trg_ws['B12'] = src_ws['B12'].value
            #distinct values only
            trg_ws['B14'] = src_ws['B14'].value
            #custom source view
            trg_ws['B16'] = src_ws['B16'].value
            #disable table prefix
            trg_ws['B18'] = src_ws['B18'].value
            #logical table name
            trg_ws['B21'] = src_ws['B21'].value
            #is dimension table
            trg_ws['B28'] = src_ws['B28'].value
            #use scd-2
            trg_ws['B29'] = src_ws['B29'].value
            #use dma
            trg_ws['B31'] = src_ws['B31'].value
            #vehicleplatformid column name
            trg_ws['B32'] = src_ws['B32'].value

        ws="Custom Source View"
        src_ws = original_src_wb[ws]
        trg_ws = trg_wb[ws]
        trg_ws['A2'] = src_ws['A2'].value

        ws="Doc"
        src_ws = original_src_wb[ws]
        trg_ws = trg_wb[ws]
        trg_ws['B2'] = src_ws['B2'].value
        trg_ws['B6'] = src_ws['B6'].value

        ws="COLUMNS"
        src_ws = original_src_wb[ws]
        trg_ws = trg_wb[ws]
        for r in range(2, 150):
            for c in range(1,15):
                #note the common o for both, could also be o+1 for one if there is an offset
                trg_ws.cell(r, c).value = src_ws.cell(r, c).value

        ws="DIMENSIONS"
        src_ws = original_src_wb[ws]
        trg_ws = trg_wb[ws]
        for r in range(4, 16):
            for c in range(1,10):
                #note the common o for both, could also be o+1 for one if there is an offset
                trg_ws.cell(r, c).value = src_ws.cell(r, c).value


        trg_wb.save(target_fullpath)
        trg_wb.close()
        original_src_wb.close()



        #finally refresh the newly create file
        Util.print(DEBUG1,"\tRefresh upgraded file ", target_fullpath)
        solution_paths.updateExcelFileWithPowershell(target_fullpath)
        Util.print(DEBUG1,"\tCopy ", target_fullpath , " to ", final_dest)
        Util.copy(target_fullpath,final_dest)

    def debugPrint(self):
        print("")
        Util.print(INFO,"File:", str(self.fullpath.name))

        Util.printLine()

class OogwayCodeBuilder:
    '''OgwayCodeBuilder'''
    #TODO: why is this limited to range(6)? Let it be more dynamic
    other_tables_step_state_machines=[[] for _ in range(6)]

    html_dim_docs=[]
    html_fact_docs=[]
    html_other_docs=[]

    def init(self, aws_account, aws_region, athena_workgroup, athena_output_location):

        self.other_tables_step_state_machines = [[] for _ in range(6)]

        self.aws_account=aws_account
        self.aws_region=aws_region
        self.athena_workgroup=athena_workgroup
        self.athena_output_location=athena_output_location

        self.steps_arn_template = stringtemplate.arn_steps_state_machine

        self.steps_arn_template=self.steps_arn_template.replace(replacetags.aws_account,aws_account)
        self.steps_arn_template=self.steps_arn_template.replace(replacetags.aws_region,aws_region)

    def copyToDeploy2(self,Dim:DvDim, source_file_name, source_path
                      , target_file_name, target_path, do_print=False):

        full_target=os.path.join(target_path,target_file_name)
        full_source=os.path.join(source_path,source_file_name)


        #print("\tCopy ", full_source, " to ", full_target)
        shutil.copy(full_source,full_target)
        with open(full_source,'r') as src_file:
            data = src_file.read()
            if(do_print):
                print(data)
            src_file.close()
            with open(full_target,'w+') as file:
                for key in DvDim.config:
                    val = Dim.getConfigValue(key);
                    replace_tag="[:" + key + "]"
                    data=data.replace(replace_tag, str(val))


                file.write(data)

    def copyToDeploy(self,
                     Dim: DvDim,
                     template_file_and_path:string,
                     solution_paths:SolutionPaths,
                     output_code_file_name:string,
                     comment:string,list_of_other_tables_arn):

        src= template_file_and_path
        dst=os.path.join(solution_paths.code_output_path,output_code_file_name)
        dst_doc=os.path.join(solution_paths.doc_output_path,"doc")

        current_date = datetime.now().strftime("%Y-%m-%d")

        #MMA
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_0']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_1']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_2']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_3']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_4']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_5']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_6']=""
        Dim.config['REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_7']=""

        for index, x in enumerate(list_of_other_tables_arn):
            val = "REPLACE_LIST_OF_OTHER_TABLES_ARN_LEVEL_" + str(index)
            Dim.config[val]=list_of_other_tables_arn[index]

        Util.print(DEBUG1,"Copy ", src, " to ", dst)
        Util.copy(src,dst)

        with open(src,'r') as src_file:
            data = src_file.read()
            src_file.close()
            with open(dst,'w+') as file:
                for key in DvDim.config:
                    replace_tag="[:" + key + "]"
                    data=data.replace(replace_tag, str(Dim.config[key]))

                file.write(data)

        return data

    def getPlatformResourceIdentifier(self,main_solution_file, resource_name:string):
        '''
        Each platform has its own resource identifiers.
        This method gets the resource indetifer template from the solution.xlsx file, replaces some tags (if any)
        and finaly returns the platform resource identifier.
        '''
        rid= main_solution_file.output_builder.platform_resource_identifier_template


        rid=rid.replace("[:REPLACE_AWS_ACCOUNT_NUMBER]",main_solution_file.aws_cloud.aws_account)
        rid=rid.replace("[:REPLACE_AWS_ACCOUNT]",main_solution_file.aws_cloud.aws_account)
        rid=rid.replace("[:REPLACE_AWS_REGION]",main_solution_file.aws_cloud.aws_region)

        rid=rid.replace("[:REPLACE_RESOURCE_IDENTIFIER]",resource_name)

        Util.print(DEBUG3,"\tResource Identifier: ", rid)
        return rid

    def create_code(self, Dim: DvDim, solution_paths:SolutionPaths, main_solution_file: DvDim):
        '''
        This method takes variables from the Excel file and injects them into the templates.
         * copies the files to the correct folder in the solution repo.
         * Some built in knowledge about some Load Methods...
         * creates some new oogway-variables and adds the to the list of available variabels (config)
        '''
        comment="LOAD"
        is_main=False
        resource_name=""

        if(Dim.table_name==""):
            is_main=True
            Dim.resource_name =main_solution_file.output_builder.solutionName + "_main"

        resource_name=Dim.resource_name
        Util.print(DEBUG1,"Resource Name:", Dim.resource_name)

        if(Dim.load_method==1): #load dimension
            if(Dim.config["REPLACE_USE_MINI_DIMENSION"]):
                Util.print(DEBUG1,"Analyzing SCD-4")
                template_file="template_1_mini_dimension"
                target_html_file =  main_solution_file.output_builder.solutionName + "_" + Dim.table_name_mini_dimension + ".html"
                target_html_file_json= main_solution_file.output_builder.solutionName + "_" + Dim.table_name_mini_dimension + "_code.html"


                full_template_file = os.path.join(solution_paths.code_template_path ,template_file + main_solution_file.output_builder.code_file_extension)

                resource_name_mini_dim=Dim.resource_name_mini_dimension

                Dim.resource_name_mini_dimension=resource_name.replace(replacetags.replace_table_name, Dim.table_name_mini_dimension )

                resource_identifier = self.getPlatformResourceIdentifier(main_solution_file, Dim.resource_name_mini_dimension)

                Dim.config["REPLACE_MINI_DIMENSION_ARN"]=resource_identifier
                Dim.config["REPLACE_PLATFORM_RESOURCE_IDENTIFIER"]=resource_identifier

                #this is code for the mini dim
                output_file=resource_name_mini_dim + main_solution_file.output_builder.code_file_extension
                Util.print(DEBUG1, "Creating code for:", output_file, ", with template:", full_template_file)
                output_code_mini_dim = self.copyToDeploy(Dim
                                                        ,full_template_file,solution_paths
                                                        ,output_file
                                                        ,comment,[])



                t="<table><th>Table</th><th>Platform Resource Identifier</th> <tr><td><a href='" + target_html_file + "'>" + Dim.table_name_mini_dimension +"</a></td><td><a href='" + target_html_file_json + "'>" + resource_identifier + "</a></td></tr></table>"
                Dim.config['REPLACE_HTML_TABLE_MINI_DIMENSION']=t
                Dim.config["REPLACE_RESOURCE_NAME"] = Dim.resource_name_mini_dimension
                Dim.config["REPLACE_OUTPUT_CODE"] = output_code_mini_dim

                #HTML Documentation

                source_html_template="template_code.html"
                self.copyToDeploy2(Dim
                                   , source_html_template
                                   ,solution_paths.documentation_templates_path
                                   ,target_html_file_json
                                   ,solution_paths.doc_output_path)

                source_html_template="template_1_mini_dimension.html"
                self.copyToDeploy2(Dim
                                   , source_html_template
                                   , solution_paths.documentation_templates_path,target_html_file
                                   , solution_paths.doc_output_path)


                Dim.config["REPLACE_RESOURCE_NAME"] = Dim.resource_name

        if Dim.load_method==2: #load fact as hive table for platform AWS
            steps_name_loop = Dim.aws_steps_loop_name
            steps_name_loop=steps_name_loop.replace(replacetags.replace_table_name, Dim.table_name )
            steps_name_loop=steps_name_loop.replace(replacetags.replace_solution_name, output_builder.solutionName )

            steps_name_jobqueue = stringtemplate.awssteps_jobqueue_state_machine_name
            steps_name_jobqueue=steps_name_jobqueue.replace(replacetags.replace_solution_name, output_builder.solutionName )


        resource_name=resource_name.replace(replacetags.replace_table_name, Dim.table_name )
        resource_name=resource_name.replace(replacetags.replace_solution_name, main_solution_file.output_builder.solutionName )

        platform_resource_identifier = self.getPlatformResourceIdentifier(main_solution_file, resource_name)

        #sets the oogway-variable REPLACE_PLATFORM_RESOURCE_IDENTIFIER
        Dim.config['REPLACE_PLATFORM_RESOURCE_IDENTIFIER'] = platform_resource_identifier


        #copy template to output dir
        template_file="template_" + str(Dim.load_method)
        full_template_file = os.path.join(solution_paths.code_template_path ,template_file + main_solution_file.output_builder.code_file_extension)
        full_template_file_loop = os.path.join(solution_paths.code_template_path, template_file + "_loop" + main_solution_file.output_builder.code_file_extension)
        full_template_file_jobqueu = os.path.join(solution_paths.code_template_path , "template_jobqueue" + main_solution_file.output_builder.code_file_extension)


        if Dim.load_method==2: #UNLOAD FACT TABLE (EXTERNAL TABLE)
            #loop is for manual loads... replace with glue???
            Dim.config["REPLACE_UNLOAD_TABLE_STEPS_NAME"] = resource_name
            self.copyToDeploy(Dim,
                              full_template_file_loop,
                              solution_paths,
                              steps_name_loop + main_solution_file.output_builder.code_file_extension,
                              comment,[])

            self.copyToDeploy(Dim,
                              full_template_file_jobqueu,
                              solution_paths,
                              steps_name_jobqueue + main_solution_file.output_builder.code_file_extension,
                              comment,[])


        #copy the actual aws steps job to deploy folder
        output_file=resource_name + main_solution_file.output_builder.code_file_extension
        Util.print(DEBUG1, "Creating code file:", output_file)
        output_code = self.copyToDeploy(Dim,
                                        full_template_file,
                                        solution_paths
                                        ,output_file
                                        ,comment,[])

        Util.print(DEBUG3,output_code)

        '''sets the oogway-variable REPLACE_OUTPUT_CODE'''
        Dim.config['REPLACE_OUTPUT_CODE']=output_code
        Dim.output_code=output_code



        #HTML Documentation
        target_html_file=""
        target_html_file_json=""
        source_html_template=template_file + ".html"

        if Dim.load_method!=0: #THIS IS MAIN ORCHESTRATOR
            Util.print(DEBUG1,"REPLACE_SOLUTION_NAME:", main_solution_file.output_builder.solutionName)

            target_html_file =  main_solution_file.output_builder.solutionName + "_" + Dim.getConfigValue("REPLACE_TARGET_TABLE_NAME") + ".html"

            Util.print(DEBUG1, "Creating html file:", target_html_file)
            data=self.copyToDeploy2(Dim,
                           source_html_template,
                           solution_paths.documentation_templates_path,
                           target_html_file,
                           solution_paths.doc_output_path,
                           )

            #create html with the actual code
            source_html_template="template_code.html"

            #target code file
            target_html_file_json = main_solution_file.output_builder.solutionName + "_" + Dim.getConfigValue("REPLACE_TARGET_TABLE_NAME") + "_code.html"

            Util.print(DEBUG1, "Creating html file:", target_html_file_json)
            self.copyToDeploy2(Dim,
                               source_html_template
                               ,solution_paths.documentation_templates_path
                               ,target_html_file_json
                               ,solution_paths.doc_output_path)

        oogway_output_code=OogwayOutputCode()
        oogway_output_code.platform_resource_identifier=platform_resource_identifier
        oogway_output_code.htnl_code_filename=target_html_file_json
        oogway_output_code.html_table_filename=target_html_file
        oogway_output_code.generated_code=output_code

        return oogway_output_code

    def prepare_jq_array_steps(self,dim:DvDim,  solution_paths: SolutionPaths, main_solution_file):

        if  main_solution_file.output_builder.orchestration!="aws_steps":
           return

        comment=""
        full_template_file = os.path.join(solution_paths.code_template_path ,"template_jobqueue_array" + main_solution_file.output_builder.code_file_extension)
        resource_name = stringtemplate.awssteps_jq_array_state_machine_name
        resource_name=resource_name.replace(replacetags.replace_solution_name, main_solution_file.output_builder.solutionName )

        output_code_file_name=resource_name + main_solution_file.output_builder.code_file_extension

        #resource_identifier= stringtemplate.arn_steps_state_machine.replace(replacetags.aws_state_machine_name,resource_name )
        platform_resource_identifier = self.getPlatformResourceIdentifier(main_solution_file, resource_name)

        #copy template to output dir

        self.copyToDeploy(dim,
                          full_template_file,
                          solution_paths,
                          output_code_file_name,
                          comment,[])

        return platform_resource_identifier

    #TODO: this only works for aws steps main state machine at the moment
    def prepare_main_script(self
                            , main_solution_file:DvDim
                            , solution_paths:SolutionPaths
                           ,  other_tables_list_of_arn
                           ,  html_table_files):

        list_of_other_tables=[[]]
        list_of_other_tables = [[] for _ in range(6)]

        output_code:OogwayOutputCode
        main_solution_file.openForBuild(main_solution_file, main_solution_file.fullpath)
        Util.print(DEBUG1, "Creating code for main",)
        output_code= self.create_code(main_solution_file,solution_paths, main_solution_file)

        comment=""
        #create a comma separated string where each element is enclosed in ""
        for index, array in enumerate(other_tables_list_of_arn):
            list=other_tables_list_of_arn[index]
            list_of_other_tables[index] = '"{0}"'.format('", "'.join(list))

        full_template_file = os.path.join(solution_paths.code_template_path
                                          , "template_0" + main_solution_file.output_builder.code_file_extension)

        resource_name = main_solution_file.output_builder.solutionName + "_main"  #stringtemplate.awssteps_main_load_state_machine_name
        output_code_file_name=resource_name + main_solution_file.output_builder.code_file_extension
        html_output_steps_name=resource_name + ".html"
        main_solution_file.config["REPLACE_MAIN_STEPS_FILENAME"] = output_code_file_name
        #resource_identifier= stringtemplate.arn_steps_state_machine.replace(replacetags.aws_state_machine_name,resource_name )
        platform_resource_identifier = self.getPlatformResourceIdentifier(main_solution_file, resource_name)

        #copy template to output dir
        main_aws_steps_json = self.copyToDeploy(main_solution_file
                                                , full_template_file,solution_paths
                                                , output_code_file_name
                                                , comment
                                                , list_of_other_tables
                                                )

        target_html_file="index.html"
        source_html_template="template_0.html"

        main_solution_file.config['REPLACE_TABLE_OF_HTML_FILES'] = html_table_files

        main_solution_file.config['REPLACE_USECASE_VIEWS_TABLE_DISPLAY']='display: none'
        main_solution_file.config['REPLACE_INCOMING_VIEWS_TABLE_DISPLAY']='display: none'

        #new: 2023-12-27 Views can now be defined in the solution file
        array = main_solution_file.getConfigValue("REPLACE_ARRAY_OF_VIEWS","")
        number_of_usecase_views=0
        if(array!=""):
            data = json.loads(array)
            for element in data:
                number_of_usecase_views=number_of_usecase_views+1
                source_view_html_file="template_0_view.html"
                target_view_html_file=element['name'] + "_view.html"

                Util.print(DEBUG3,target_view_html_file)

                main_solution_file.config["REPLACE_VIEW_NAME"] = element['name']
                main_solution_file.config["REPLACE_CREATE_VIEW"] = element['q']

                self.copyToDeploy2(main_solution_file,
                            source_view_html_file,
                            solution_paths.documentation_templates_path,
                            target_view_html_file,
                            solution_paths.doc_output_path,
                            )

        if(number_of_usecase_views>0):
             main_solution_file.config['REPLACE_USECASE_VIEWS_TABLE_DISPLAY']='display: block'

        array=""
        array = main_solution_file.getConfigValue("REPLACE_ARRAY_OF_INCOMING_VIEWS","[]")

        number_of_temp_views=0
        if(array!=""):

            data = json.loads(array)
            for element in data:
                number_of_temp_views=number_of_temp_views+1
                source_view_html_file="template_0_view.html"
                target_view_html_file=element['name'] + "_view.html"

                Util.print(DEBUG3, target_view_html_file)

                main_solution_file.config["REPLACE_VIEW_NAME"] = element['name']
                main_solution_file.config["REPLACE_CREATE_VIEW"] = element['q']
                self.copyToDeploy2(main_solution_file,
                            source_view_html_file,
                            solution_paths.documentation_templates_path,
                            target_view_html_file,
                            solution_paths.doc_output_path,
                            )
        if(number_of_temp_views>0):
             main_solution_file.config['REPLACE_INCOMING_VIEWS_TABLE_DISPLAY']='display: block'


        #MAIN DOCUMENTATION
        main_solution_file.config["REPLACE_OUTPUT_CODE"] = main_aws_steps_json
        main_solution_file.config["REPLACE_MAIN_STEPS_HTML_FILENAME"]=html_output_steps_name

        self.copyToDeploy2(main_solution_file,
                           source_html_template,
                           solution_paths.documentation_templates_path,
                           target_html_file,
                           solution_paths.doc_output_path,
                           )

        source_html_template="template_templates.html"
        target_html_file="templates.html"

        self.copyToDeploy2(main_solution_file,
                           source_html_template,
                           solution_paths.documentation_templates_path,
                           target_html_file,
                           solution_paths.doc_output_path,
                           )


        #MAIN CODE
        self.copyToDeploy2(main_solution_file,
                            "template_code.html",
                            solution_paths.documentation_templates_path,
                            html_output_steps_name,
                            solution_paths.doc_output_path
                            )
        main_solution_file.platform_resource_identifier = platform_resource_identifier
        return main_solution_file

    #MAIN FLOW METHODS
    def create_main_aws_state_machines_old(self,solution_name, output_path, output_builder: OutputBuilder ):
        #create final state machine that will run all of the others in a queue with x number of simultanious executions.

        aws_initil_load_main = self._get_aws_steps_load_main(self.load_step_state_machines, "Initial load")
        self.save_json(output_path, solution_name +  "_main_steps" + output_builder.code_file_extension , aws_initil_load_main)

        aws_incremental_load_main = self._get_aws_steps_load_main(self.incremental_load_step_state_machines, "Incremental load")
        self.save_json(output_path, solution_name +  "_main_incremental_steps" + output_builder.code_file_extension , aws_incremental_load_main)

    #PRIVATE METHODS
    def save_json(self,p,f, content):

        fp=os.path.join(p, f)
        f2= open(fp,"w")
        s = json.dumps(content)
        f2.write(s)




